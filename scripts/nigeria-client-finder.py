#!/usr/bin/env python3
"""
尼日利亚客户开发自动化搜索脚本
适用行业：家装五金、门锁、拉手、导轨、铰链、家具配件等出口贸易

⚠️ 合规说明：
- 不使用爬虫抓取社交媒体（违反 ToS 会封号）
- Google 搜索使用官方 API 或手动辅助
- 社交媒体生成精准搜索链接，人工查看联系

作者：为晓雷哥哥定制
日期：2026-04-08
"""

import csv
import json
import time
import re
import urllib.parse
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional
from dataclasses import dataclass, asdict

# ============ 配置区域 ============

CONFIG = {
    # 行业关键词（中英文）
    'product_keywords': [
        'door lock', 'door handle', 'cabinet hinge', 'drawer slide', 'furniture hardware',
        'edge banding', 'furniture glue', 'cabinet pull', 'hinge', 'furniture accessories',
        '门锁', '门把手', '铰链', '导轨', '家具五金', '封边条', '家具配件'
    ],
    
    # 客户类型关键词
    'buyer_types': [
        'importer', 'wholesaler', 'distributor', 'dealer', 'trading company',
        'building materials', 'hardware store', 'construction supplier',
        '进口商', '批发商', '经销商'
    ],
    
    # 目标市场
    'target_country': 'Nigeria',
    'target_cities': ['Lagos', 'Abuja', 'Kano', 'Port Harcourt', 'Ibadan'],
    
    # 输出文件（相对于脚本所在目录）
    'output_csv': 'nigeria_clients.csv',
    'output_excel': 'nigeria_clients.xlsx',
    'output_links': 'nigeria_search_links.json',
    'output_template': 'client_template.csv',
    
    # Google Custom Search API（可选，有免费额度）
    # 获取方式：https://developers.google.com/custom-search/v1/overview
    'google_api_key': '',  # 填入你的 API Key
    'google_cx': '',       # 填入你的 Search Engine ID
}

# ============ 数据结构 ============

@dataclass
class ClientInfo:
    """客户信息数据结构"""
    company_name: str = ''
    contact_person: str = ''
    email: str = ''
    phone: str = ''
    website: str = ''
    address: str = ''
    city: str = ''
    social_links: Dict[str, str] = None
    source: str = ''  # 来源：Google/Facebook/LinkedIn 等
    search_keywords: str = ''
    notes: str = ''
    found_date: str = ''
    
    def __post_init__(self):
        if self.social_links is None:
            self.social_links = {}
        if not self.found_date:
            self.found_date = datetime.now().strftime('%Y-%m-%d %H:%M')

# ============ 关键词组合生成 ============

def generate_search_queries() -> List[str]:
    """
    生成组合搜索关键词
    
    返回：
        搜索关键词列表
    """
    queries = []
    
    # 组合 1: 产品 + 买家类型 + 国家
    for product in CONFIG['product_keywords'][:5]:  # 限制数量避免过多
        for buyer_type in CONFIG['buyer_types'][:4]:
            query = f'"{product}" {buyer_type} in {CONFIG["target_country"]}'
            queries.append(query)
    
    # 组合 2: 产品 + 城市
    for product in CONFIG['product_keywords'][:3]:
        for city in CONFIG['target_cities'][:3]:
            query = f'"{product}" supplier distributor "{city}" Nigeria'
            queries.append(query)
    
    # 组合 3: 行业目录搜索
    queries.extend([
        'Nigeria hardware importers directory',
        'Nigeria building materials suppliers list',
        'Nigeria furniture hardware buyers database',
        'Lagos hardware wholesalers contact',
        'Nigeria construction materials importers email'
    ])
    
    # 组合 4: 社交媒体搜索
    queries.extend([
        'site:facebook.com "hardware" Nigeria importer',
        'site:linkedin.com/company "building materials" Nigeria',
        'site:instagram.com "furniture hardware" Nigeria'
    ])
    
    return list(set(queries))  # 去重

# ============ Google 搜索（API 方式） ============

def google_search_api(query: str, num_results: int = 5) -> List[Dict]:
    """
    使用 Google Custom Search API 搜索（合规方式）
    
    参数:
        query: 搜索关键词
        num_results: 返回结果数量
    
    返回:
        搜索结果列表
    """
    import urllib.request
    
    if not CONFIG['google_api_key'] or not CONFIG['google_cx']:
        print(f"⚠️  未配置 Google API，跳过：{query[:50]}...")
        return []
    
    try:
        url = (
            f"https://www.googleapis.com/customsearch/v1?"
            f"key={CONFIG['google_api_key']}&"
            f"cx={CONFIG['google_cx']}&"
            f"q={urllib.parse.quote(query)}&"
            f"num={num_results}"
        )
        
        with urllib.request.urlopen(url, timeout=10) as response:
            data = json.loads(response.read().decode())
        
        results = []
        for item in data.get('items', []):
            results.append({
                'title': item.get('title', ''),
                'link': item.get('link', ''),
                'snippet': item.get('snippet', ''),
                'display_link': item.get('displayLink', '')
            })
        
        time.sleep(0.5)  # 避免频率限制
        return results
        
    except Exception as e:
        print(f"❌ Google API 错误：{e}")
        return []

# ============ Google 搜索（浏览器辅助方式） ============

def generate_google_search_url(query: str) -> str:
    """
    生成 Google 搜索 URL（手动搜索用）
    
    参数:
        query: 搜索关键词
    
    返回:
        Google 搜索 URL
    """
    import urllib.parse
    return f"https://www.google.com/search?q={urllib.parse.quote(query)}"

def generate_social_search_urls(product: str, city: str = '') -> Dict[str, str]:
    """
    生成各社交媒体精准搜索链接
    
    参数:
        product: 产品关键词
        city: 城市名（可选）
    
    返回:
        各平台搜索链接字典
    """
    location = f" {city} Nigeria" if city else " Nigeria"
    
    return {
        'Google': generate_google_search_url(f'"{product}" importer wholesaler{location}'),
        'Facebook': f"https://www.facebook.com/search/posts/?q={urllib.parse.quote(product + ' hardware Nigeria')}",
        'Facebook_Companies': f"https://www.facebook.com/search/pages/?q={urllib.parse.quote('building materials Nigeria')}",
        'LinkedIn': f"https://www.linkedin.com/search/results/companies/?keywords={urllib.parse.quote(product + ' Nigeria')}&origin=GLOBAL_SEARCH_HEADER",
        'LinkedIn_People': f"https://www.linkedin.com/search/results/people/?keywords={urllib.parse.quote(product + ' buyer Nigeria')}",
        'Instagram': f"https://www.instagram.com/explore/tags/{product.replace(' ', '')}nigeria/",
        'Google_Maps': f"https://www.google.com/maps/search/{urllib.parse.quote(product + ' store ' + (city or 'Lagos') + ' Nigeria')}",
    }

# ============ 信息提取 ============

def extract_email(text: str) -> List[str]:
    """从文本中提取邮箱"""
    pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    return list(set(re.findall(pattern, text, re.IGNORECASE)))

def extract_phone(text: str, country_code: str = '+234') -> List[str]:
    """从文本中提取电话号码（尼日利亚）"""
    patterns = [
        r'\+234\s?\d{3}\s?\d{3}\s?\d{4}',
        r'\+234\d{10}',
        r'234\d{10}',
        r'0\d{3}\s?\d{3}\s?\d{4}',
        r'\+\d{1,3}\s?\(?\d{2,4}\)?\s?\d{3,4}\s?\d{4}'
    ]
    phones = []
    for pattern in patterns:
        phones.extend(re.findall(pattern, text))
    return list(set(phones))

def extract_company_name(text: str, url: str = '') -> str:
    """从文本或 URL 中提取公司名"""
    # 常见公司后缀
    suffixes = [
        r'\s+Ltd\.?', r'\s+Limited', r'\s+LLC', r'\s+Inc\.?',
        r'\s+Co\.?', r'\s+Company', r'\s+Enterprise', r'\s+Trading'
    ]
    
    # 从 URL 提取
    if url:
        domain = url.replace('www.', '').split('/')[0].replace('.com', '').replace('.ng', '')
        domain = domain.replace('-', ' ').replace('.', ' ').title()
        if domain and len(domain) > 3:
            return domain
    
    # 从标题提取
    if text:
        for suffix in suffixes:
            match = re.search(r'([^|]+)' + suffix, text, re.IGNORECASE)
            if match:
                return match.group(1).strip()
    
    return ''

# ============ 结果保存 ============

def save_to_csv(clients: List[ClientInfo], filename: str):
    """保存结果到 CSV"""
    if not clients:
        print("⚠️  没有数据可保存")
        return
    
    # 使用脚本所在目录
    script_dir = Path(__file__).parent
    output_path = script_dir / filename
    
    fieldnames = [
        'company_name', 'contact_person', 'email', 'phone', 'website',
        'address', 'city', 'facebook', 'linkedin', 'instagram',
        'source', 'search_keywords', 'notes', 'found_date'
    ]
    
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        
        for client in clients:
            row = asdict(client)
            row['facebook'] = client.social_links.get('facebook', '')
            row['linkedin'] = client.social_links.get('linkedin', '')
            row['instagram'] = client.social_links.get('instagram', '')
            writer.writerow(row)
    
    print(f"✅ 已保存到：{output_path}")

def save_to_excel(clients: List[ClientInfo], filename: str):
    """保存结果到 Excel（需要 openpyxl）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # 使用脚本所在目录
        script_dir = Path(__file__).parent
        output_path = script_dir / filename
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Nigeria Clients"
        
        # 表头
        headers = [
            '公司名', '联系人', '邮箱', '电话', '网站',
            '地址', '城市', 'Facebook', 'LinkedIn', 'Instagram',
            '来源', '搜索关键词', '备注', '发现日期'
        ]
        ws.append(headers)
        
        # 表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 数据行
        for client in clients:
            ws.append([
                client.company_name,
                client.contact_person,
                client.email,
                client.phone,
                client.website,
                client.address,
                client.city,
                client.social_links.get('facebook', ''),
                client.social_links.get('linkedin', ''),
                client.social_links.get('instagram', ''),
                client.source,
                client.search_keywords,
                client.notes,
                client.found_date
            ])
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        wb.save(output_path)
        print(f"✅ 已保存到：{output_path}")
        
    except ImportError:
        print("⚠️  未安装 openpyxl，跳过 Excel 保存。运行：pip install openpyxl")

# ============ 主流程 ============

def manual_search_mode():
    """
    手动搜索辅助模式
    生成搜索链接和关键词，人工查看并记录
    """
    print("\n" + "="*60)
    print("🔍 尼日利亚客户开发 - 手动搜索辅助模式")
    print("="*60)
    
    # 获取脚本所在目录作为输出目录
    script_dir = Path(__file__).parent
    
    # 生成搜索链接
    search_links = {}
    for product in CONFIG['product_keywords'][:5]:
        links = generate_social_search_urls(product, 'Lagos')
        search_links[product] = links
    
    # 输出搜索指南
    guide = f"""
📋 搜索指南（按顺序执行）

【第一步】Google 搜索（找官网和邮箱）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
    for i, query in enumerate(generate_search_queries()[:10], 1):
        url = generate_google_search_url(query)
        guide += f"{i}. {query[:60]}...\n   🔗 {url}\n\n"
    
    guide += """
【第二步】LinkedIn 搜索（找公司和采购负责人）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. 访问：https://www.linkedin.com/
2. 搜索关键词：
   - "building materials Nigeria"
   - "hardware importer Lagos"
   - "furniture accessories Nigeria"
3. 筛选：Companies → 地点：Nigeria
4. 查看公司页面，找：
   - About 里的官网和邮箱
   - People 里的采购负责人

【第三步】Facebook 搜索（找中小批发商）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. 访问：https://www.facebook.com/
2. 搜索：
   - "hardware store Lagos"
   - "building materials Nigeria"
   - "furniture fittings Abuja"
3. 筛选：Pages → 地点：Nigeria
4. 查看 Page Info 里的联系方式

【第四步】Instagram 搜索（找零售商）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. 访问：https://www.instagram.com/
2. 搜索标签：
   - #nigeriahardware
   - #lagosbuildingmaterials
   - #furniturenigeria
3. 查看主页 Bio 里的联系方式

【第五步】Google Maps（找实体店）
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
1. 访问：https://www.google.com/maps
2. 搜索：
   - "hardware store Lagos"
   - "building materials Ikeja"
   - "furniture accessories Nigeria"
3. 查看商家信息里的电话和网站

【第六步】行业目录网站
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
推荐网站：
• https://www.nigeriainfo.fm/lagos/business/
• https://www.vconnect.com/
• https://www.finelib.com/
• https://www.tradekey.com/nigeria/
• https://www.go4worldbusiness.com/buyers/nigeria/

"""
    
    print(guide)
    
    # 保存搜索链接到文件
    links_file = script_dir / CONFIG.get('output_links', 'nigeria_search_links.json')
    with open(links_file, 'w', encoding='utf-8') as f:
        json.dump(search_links, f, indent=2, ensure_ascii=False)
    print(f"✅ 搜索链接已保存：{links_file}")
    
    # 创建空白客户记录模板
    template_file = script_dir / CONFIG.get('output_template', 'client_template.csv')
    with open(template_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow([
            '公司名', '联系人', '邮箱', '电话', '网站',
            '地址', '城市', 'Facebook', 'LinkedIn', 'Instagram',
            '来源', '搜索关键词', '备注', '发现日期'
        ])
        # 添加 3 行空白模板
        for _ in range(3):
            writer.writerow([''] * 14)
    print(f"✅ 客户记录模板已创建：{template_file}")
    
    return search_links

def interactive_entry_mode():
    """
    交互式录入模式
    手动搜索后，在此录入客户信息
    """
    print("\n" + "="*60)
    print("📝 客户信息录入模式")
    print("="*60)
    
    clients = []
    
    while True:
        print(f"\n【录入第 {len(clients)+1} 个客户】(输入 q 结束)")
        
        company = input("公司名：").strip()
        if company.lower() == 'q':
            break
        
        client = ClientInfo(
            company_name=company,
            contact_person=input("联系人：").strip(),
            email=input("邮箱：").strip(),
            phone=input("电话：").strip(),
            website=input("网站：").strip(),
            address=input("地址：").strip(),
            city=input("城市：").strip(),
            source=input("来源 (Google/FB/LinkedIn 等)：").strip(),
            search_keywords=input("搜索关键词：").strip(),
            notes=input("备注：").strip()
        )
        
        # 社交媒体
        fb = input("Facebook: ").strip()
        if fb:
            client.social_links['facebook'] = fb
        li = input("LinkedIn: ").strip()
        if li:
            client.social_links['linkedin'] = li
        ig = input("Instagram: ").strip()
        if ig:
            client.social_links['instagram'] = ig
        
        clients.append(client)
        print(f"✅ 已记录：{company}")
    
    if clients:
        save_to_csv(clients, CONFIG['output_csv'])
        save_to_excel(clients, CONFIG['output_excel'])
    
    return clients

# ============ 主函数 ============

def quick_start_mode():
    """
    快速启动模式 - 无需交互，直接生成搜索链接和指南
    """
    print("\n🚀 快速生成搜索链接...\n")
    manual_search_mode()
    print("\n" + "="*60)
    print("✅ 完成！请查看以下文件：")
    print("   📄 nigeria_search_links.json - 搜索链接")
    print("   📄 client_template.csv - 客户记录模板")
    print("   📖 README-尼日利亚客户开发.md - 详细使用说明")
    print("="*60)

def main():
    """主程序入口"""
    print("""
╔══════════════════════════════════════════════════════════╗
║     尼日利亚客户开发自动化搜索脚本                        ║
║     家装五金 · 门锁 · 拉手 · 导轨 · 铰链 · 家具配件       ║
╚══════════════════════════════════════════════════════════╝
    """)
    
    # 检查命令行参数
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == '--quick':
        quick_start_mode()
        return
    
    print("请选择模式：")
    print("1. 🔍 手动搜索辅助模式（生成搜索链接和指南）")
    print("2. 📝 交互式录入模式（搜索后录入客户信息）")
    print("3. 🤖 API 自动搜索模式（需配置 Google API）")
    print("4. 🚀 快速启动模式（无需交互，直接生成）")
    print()
    print("💡 提示：也可运行 python3 nigeria-client-finder.py --quick 直接生成")
    print()
    
    try:
        choice = input("输入数字选择 [1/2/3/4]: ").strip()
    except EOFError:
        # 非交互模式下自动使用快速启动
        quick_start_mode()
        return
    
    if choice == '1':
        manual_search_mode()
        print("\n💡 提示：搜索后使用模式 2 录入客户信息")
    elif choice == '2':
        interactive_entry_mode()
    elif choice == '3':
        if not CONFIG['google_api_key'] or not CONFIG['google_cx']:
            print("\n❌ 未配置 Google API Key 和 CX")
            print("📖 获取方式：https://developers.google.com/custom-search/v1/overview")
            print("💡 或先在脚本中配置 CONFIG['google_api_key'] 和 CONFIG['google_cx']")
        else:
            print("\n🔍 开始自动搜索...")
            # 可在此扩展 API 搜索逻辑
            manual_search_mode()
    elif choice == '4':
        quick_start_mode()
    else:
        print("无效选择")

if __name__ == '__main__':
    main()
